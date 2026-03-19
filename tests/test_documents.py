"""Testes para tools de documentos (Word e Excel)."""

from __future__ import annotations

import pytest

from denai.tools.documents import create_document, create_spreadsheet


@pytest.mark.asyncio
async def test_create_document_basic(tmp_path):
    """Cria um documento Word simples."""
    path = str(tmp_path / "test.docx")
    result = await create_document({
        "path": path,
        "content": [
            {"type": "heading", "text": "Título", "level": 1},
            {"type": "paragraph", "text": "Olá mundo!"},
        ],
    })
    assert "✅" in result
    assert "test.docx" in result
    assert (tmp_path / "test.docx").exists()


@pytest.mark.asyncio
async def test_create_document_with_table(tmp_path):
    """Cria documento com tabela."""
    path = str(tmp_path / "tabela.docx")
    result = await create_document({
        "path": path,
        "content": [
            {"type": "heading", "text": "Relatório", "level": 1},
            {
                "type": "table",
                "headers": ["Nome", "Idade", "Cidade"],
                "rows": [
                    ["Maria", "30", "SP"],
                    ["João", "25", "RJ"],
                ],
            },
        ],
    })
    assert "✅" in result
    assert (tmp_path / "tabela.docx").exists()


@pytest.mark.asyncio
async def test_create_document_with_bullets(tmp_path):
    """Cria documento com lista de bullets."""
    path = str(tmp_path / "lista.docx")
    result = await create_document({
        "path": path,
        "content": [
            {"type": "heading", "text": "Compras", "level": 2},
            {"type": "bullet_list", "items": ["Arroz", "Feijão", "Macarrão"]},
        ],
    })
    assert "✅" in result
    assert (tmp_path / "lista.docx").exists()


@pytest.mark.asyncio
async def test_create_document_auto_extension(tmp_path):
    """Adiciona .docx se não fornecido."""
    path = str(tmp_path / "sem_ext")
    result = await create_document({
        "path": path,
        "content": [{"type": "paragraph", "text": "teste"}],
    })
    assert "✅" in result
    assert (tmp_path / "sem_ext.docx").exists()


@pytest.mark.asyncio
async def test_create_document_missing_path():
    """Erro quando path não é fornecido."""
    result = await create_document({"content": []})
    assert "❌" in result


@pytest.mark.asyncio
async def test_create_document_string_content(tmp_path):
    """Cria documento com conteúdo string simples."""
    path = str(tmp_path / "simples.docx")
    result = await create_document({
        "path": path,
        "content": "Texto simples em string.",
    })
    assert "✅" in result
    assert (tmp_path / "simples.docx").exists()


@pytest.mark.asyncio
async def test_create_spreadsheet_basic(tmp_path):
    """Cria planilha Excel simples."""
    path = str(tmp_path / "dados.xlsx")
    result = await create_spreadsheet({
        "path": path,
        "sheets": [
            {
                "name": "Vendas",
                "headers": ["Produto", "Qtd", "Preço"],
                "rows": [
                    ["Arroz", "10", "5.50"],
                    ["Feijão", "5", "8.90"],
                ],
            },
        ],
    })
    assert "✅" in result
    assert "dados.xlsx" in result
    assert "2 linhas" in result
    assert (tmp_path / "dados.xlsx").exists()


@pytest.mark.asyncio
async def test_create_spreadsheet_multiple_sheets(tmp_path):
    """Cria planilha com múltiplas abas."""
    path = str(tmp_path / "multi.xlsx")
    result = await create_spreadsheet({
        "path": path,
        "sheets": [
            {
                "name": "Receitas",
                "headers": ["Mês", "Valor"],
                "rows": [["Jan", "1000"], ["Fev", "1200"]],
            },
            {
                "name": "Despesas",
                "headers": ["Mês", "Valor"],
                "rows": [["Jan", "800"], ["Fev", "900"]],
            },
        ],
    })
    assert "✅" in result
    assert "2 aba(s)" in result
    assert (tmp_path / "multi.xlsx").exists()


@pytest.mark.asyncio
async def test_create_spreadsheet_auto_extension(tmp_path):
    """Adiciona .xlsx se não fornecido."""
    path = str(tmp_path / "sem_ext")
    result = await create_spreadsheet({
        "path": path,
        "sheets": [{"name": "A", "headers": ["X"], "rows": [["1"]]}],
    })
    assert "✅" in result
    assert (tmp_path / "sem_ext.xlsx").exists()


@pytest.mark.asyncio
async def test_create_spreadsheet_missing_path():
    """Erro quando path não é fornecido."""
    result = await create_spreadsheet({"sheets": []})
    assert "❌" in result


@pytest.mark.asyncio
async def test_create_spreadsheet_empty_sheets():
    """Erro quando sheets é vazio."""
    result = await create_spreadsheet({"path": "/tmp/x.xlsx", "sheets": []})
    assert "❌" in result


@pytest.mark.asyncio
async def test_create_spreadsheet_numeric_conversion(tmp_path):
    """Números em string são convertidos para numérico no Excel."""
    path = str(tmp_path / "numeros.xlsx")
    result = await create_spreadsheet({
        "path": path,
        "sheets": [
            {
                "name": "Dados",
                "headers": ["Int", "Float", "Texto"],
                "rows": [["42", "3.14", "abc"]],
            },
        ],
    })
    assert "✅" in result

    # Verifica que os valores foram convertidos
    from openpyxl import load_workbook
    wb = load_workbook(str(tmp_path / "numeros.xlsx"))
    ws = wb.active
    assert ws.cell(row=2, column=1).value == 42
    assert ws.cell(row=2, column=2).value == 3.14
    assert ws.cell(row=2, column=3).value == "abc"
